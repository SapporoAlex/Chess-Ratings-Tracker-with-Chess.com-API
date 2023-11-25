import openpyxl as xl
from openpyxl.chart import Reference, LineChart
from chessdotcom import Client, get_player_stats
from datetime import datetime

# Retrieving ratings data from chessdotcom
Client.request_config['headers']['User-Agent'] = 'My Python Application. Contact me at email@example.com'
data = get_player_stats('sapporoalex', tts=0).json
last_rating_chess_daily = data['stats']['chess_daily']['last']['rating']
last_rating_chess_rapid = data['stats']['chess_rapid']['last']['rating']
last_rating_chess_blitz = data['stats']['chess_blitz']['last']['rating']

# Setting the year, month and day
current_datetime = datetime.now()
current_day = current_datetime.day
current_month = current_datetime.month
current_year = current_datetime.year
current_date = current_datetime.date()

# For testing
"""
print("Last rating in chess daily:", last_rating_chess_daily)
print("Last rating in chess rapid:", last_rating_chess_rapid)
print("Last rating in chess blitz:", last_rating_chess_blitz)
print(current_year)
print(current_month)
print(current_day)
"""

# Dictionary so the month can be matched to the correct Excel sheet
months_dict = {
    1: 'Jan',
    2: 'Feb',
    3: 'Mar',
    4: 'Apr',
    5: 'May',
    6: 'Jun',
    7: 'Jul',
    8: 'Aug',
    9: 'Sept',
    10: 'Oct',
    11: 'Nov',
    12: 'Dec'
}

# Loading workbook
chess_ratings = xl.load_workbook(f'Chess Ratings {current_year}.xlsx')

# Locating the current month's sheet using the dictionary
this_month_sheet = chess_ratings[months_dict[current_month]]

# Assigning cells to variables
date_title = this_month_sheet['a1']
daily_title = this_month_sheet['b1']
rapid_title = this_month_sheet['c1']
blitz_title = this_month_sheet['d1']
date_cell = this_month_sheet.cell(current_day + 1, 1)
daily_cell = this_month_sheet.cell(current_day + 1, 2)
rapid_cell = this_month_sheet.cell(current_day + 1, 3)
blitz_cell = this_month_sheet.cell(current_day + 1, 4)

# Assigning values to cells
date_title.value = 'Date'
daily_title.value = 'Daily'
rapid_title.value = 'Rapid'
blitz_title.value = 'Blitz'
date_cell.value = current_date
daily_cell.value = last_rating_chess_daily
rapid_cell.value = last_rating_chess_rapid
blitz_cell.value = last_rating_chess_blitz

# Elo line charts
current_month_sheet = chess_ratings[months_dict[current_month]]
daily_values = Reference(current_month_sheet,
                         min_row=2,
                         max_row=current_day + 1,
                         min_col=2,
                         max_col=2)

chart_daily = LineChart()
chart_daily.title = "Daily Chess"
chart_daily.y_axis.title = "Rating"
chart_daily.x_axis.title = "Day"
chart_daily.add_data(daily_values)
current_month_sheet.add_chart(chart_daily, 'e1')

rapid_values = Reference(current_month_sheet,
                         min_row=2,
                         max_row=current_day + 1,
                         min_col=3,
                         max_col=3)

chart_rapid = LineChart()
chart_rapid.title = "Rapid Chess"
chart_rapid.y_axis.title = "Rating"
chart_rapid.x_axis.title = "Day"
chart_rapid.add_data(rapid_values)
current_month_sheet.add_chart(chart_rapid, 'e13')

blitz_values = Reference(current_month_sheet,
                         min_row=2,
                         max_row=current_day + 1,
                         min_col=4,
                         max_col=4)

chart_blitz = LineChart()
chart_blitz.title = "Blitz Chess"
chart_blitz.y_axis.title = "Rating"
chart_blitz.x_axis.title = "Day"
chart_blitz.add_data(blitz_values)
current_month_sheet.add_chart(chart_blitz, 'e25')

# Updating the overview page on the first of every month
if current_day == 1:
    overview_sheet = chess_ratings['Overview']
    month_title = overview_sheet['a1']
    overview_daily_title = overview_sheet['b1']
    overview_rapid_title = overview_sheet['c1']
    overview_blitz_title = overview_sheet['d1']
    overview_month_cell = overview_sheet.cell(current_month + 1, 1)
    overview_daily_cell = overview_sheet.cell(current_month + 1, 2)
    overview_rapid_cell = overview_sheet.cell(current_month + 1, 3)
    overview_blitz_cell = overview_sheet.cell(current_month + 1, 4)

    month_title.value = 'Month'
    overview_daily_title.value = 'Daily'
    overview_rapid_title.value = 'Rapid'
    overview_blitz_title.value = 'Blitz'
    overview_month_cell.value = months_dict[current_month]
    overview_daily_cell.value = current_month_sheet['b2']
    overview_rapid_cell.value = current_month_sheet['c2']
    overview_blitz_cell.value = current_month_sheet['d2']

    daily_values = Reference(overview_sheet,
                             min_row=2,
                             max_row=current_month + 1,
                             min_col=2,
                             max_col=2)

    chart_daily = LineChart()
    chart_daily.title = "Daily Chess"
    chart_daily.y_axis.title = "Rating"
    chart_daily.x_axis.title = "Day"
    chart_daily.add_data(daily_values)
    current_month_sheet.add_chart(chart_daily, 'e1')

    rapid_values = Reference(overview_sheet,
                             min_row=2,
                             max_row=current_month + 1,
                             min_col=3,
                             max_col=3)

    chart_rapid = LineChart()
    chart_rapid.title = "Rapid Chess"
    chart_rapid.y_axis.title = "Rating"
    chart_rapid.x_axis.title = "Day"
    chart_rapid.add_data(rapid_values)
    current_month_sheet.add_chart(chart_rapid, 'e13')

    blitz_values = Reference(overview_sheet,
                             min_row=2,
                             max_row=current_month + 1,
                             min_col=4,
                             max_col=4)

    chart_blitz = LineChart()
    chart_blitz.title = "Blitz Chess"
    chart_blitz.y_axis.title = "Rating"
    chart_blitz.x_axis.title = "Day"
    chart_blitz.add_data(blitz_values)
    current_month_sheet.add_chart(chart_blitz, 'e25')

chess_ratings.save(f'Chess Ratings {current_year}.xlsx')
